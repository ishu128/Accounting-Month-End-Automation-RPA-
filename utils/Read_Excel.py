# ============================================================
# READ PROPERTY DATA - AUTO CHECK DESKTOP VERSION
# ============================================================


import os
import time
import logging
import win32com.client as win32
from openpyxl import load_workbook
from datetime import datetime, timedelta

log = logging.getLogger(__name__)

def read_property_data():
    """
    Checks for 'AME Try.xlsm' in:
    1️⃣ Normal Desktop
    2️⃣ OneDrive Desktop

    Reads directly using openpyxl.
    No Excel COM. No SaveCopyAs.
    """

    log.info("🔍 Searching for AME Try.xlsm on Desktop locations...")

    user_home = os.path.expanduser("~")

    desktop_normal = os.path.join(user_home, "Desktop")
    desktop_onedrive = os.path.join(user_home, "OneDrive - Greystar", "Desktop")

    file_name = "AME Try.xlsm"

    possible_paths = [
        os.path.join(desktop_normal, file_name),
        os.path.join(desktop_onedrive, file_name)
    ]

    file_path = None

    for path in possible_paths:
        if os.path.exists(path):
            file_path = path
            break

    if not file_path:
        raise FileNotFoundError(
            "❌ 'AME Try.xlsm' not found in:\n"
            f"• {possible_paths[0]}\n"
            f"• {possible_paths[1]}"
        )

    log.info(f"✅ Found file at: {file_path}")

    # ------------------------------------------------------------
    # Read directly using openpyxl
    # ------------------------------------------------------------
    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb["Properties"]

    properties = []
    today = datetime.today()

    for row in ws.iter_rows(min_row=2, values_only=True):

        prop_name = str(row[0]).strip() if row[0] else ""
        prop_code = str(row[1]).strip() if row[1] else ""
        pms_type = str(row[3]).strip() if row[3] else ""
        batch_option = str(row[4]).strip().lower() if row[4] else "no"
        save_path = str(row[5]).strip() if row[5] else ""

        excel_last_ame = row[6]
        excel_current_ame = row[7]

        # Last AME Date
        if not excel_last_ame:
            first_day_this_month = today.replace(day=1)
            last_month_last_day = first_day_this_month - timedelta(days=1)
            last_ame_date = last_month_last_day.replace(day=1).strftime("%m/%d/%Y")
        else:
            last_ame_date = (
                excel_last_ame.strftime("%m/%d/%Y")
                if isinstance(excel_last_ame, datetime)
                else str(excel_last_ame).strip()
            )

        # Current AME Date
        if not excel_current_ame:
            current_ame_date = today.strftime("%m/%d/%Y")
        else:
            current_ame_date = (
                excel_current_ame.strftime("%m/%d/%Y")
                if isinstance(excel_current_ame, datetime)
                else str(excel_current_ame).strip()
            )

        if prop_name and prop_code and pms_type.lower() == "yardi":

            properties.append({
                "name": prop_name,
                "code": prop_code,
                "pms": pms_type,
                "batch_option": batch_option,
                "save_path": save_path,
                "last_ame_date": last_ame_date,
                "current_ame_date": current_ame_date
            })

    wb.close()

    log.info(f"✅ Loaded {len(properties)} Yardi properties.")

    return properties
